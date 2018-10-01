# -*- coding: utf-8 -*-

import sys
import os
import io
import json
import logging
import time
import datetime
import re

from openpyxl import load_workbook
from openpyxl import Workbook

class PolicyRuleCompare:

    workbook = None

    logging.basicConfig(filename='policyRuleCompare.log', level=logging.INFO)

    def __init__(self, cmgfile, ngbasefile, ngrulefile, ngfilterfile, resultfile):
        try:
            self.cmgfile = load_workbook(filename = cmgfile, read_only=True)
            self.ngbasefile = load_workbook(filename = ngbasefile, read_only=True)
            self.ngrulefile = load_workbook(filename = ngrulefile, read_only=True)
            self.ngfilterfile = load_workbook(filename = ngfilterfile, read_only=True)
            self.resultfilename = resultfile
            self.resultfile = Workbook(write_only=True)
            self.cmgdata = {}
            self.ngdata = {}
        except Exception as e:
            print e

    def filldata(self, wb, sheetname, datastruct, index, cols):
        ws = wb[sheetname]
        datas = []
        for row in ws.rows:
            line = []
            col = 0
            if row[index].value == None:
                break
            for cell in row:
                line.append(cell.value)
                col = col + 1
                if (col == cols):
                    break
            datas.append(line)
        datastruct[sheetname] = datas

    def init(self):
        self.filldata(self.cmgfile, 'PRB', self.cmgdata, 0, 2)
        self.filldata(self.cmgfile, 'PR', self.cmgdata, 0, 7)
        self.filldata(self.cmgfile, 'CRU', self.cmgdata, 0, 6)
        self.filldata(self.cmgfile, 'PRU', self.cmgdata, 0, 8)
        #print('cmgdata: ', self.cmgdata)
        self.filldata(self.ngbasefile, 'base', self.ngdata, 0, 5)
        self.filldata(self.ngrulefile, 'rule', self.ngdata, 0, 21)
        self.filldata(self.ngfilterfile, 'filter', self.ngdata, 4, 25)
        #print('ngdata: ', self.ngdata)

    def make_return(self, resultcode, result):
        return {'resultcode': resultcode, 'result': result}

    def excel_fill(self):
        try:
            pass     
        except Exception as e:
            return self.make_return(0, "Error PM_Excelfill: %s" % e)
        
    
    def prn_obj(self, obj):
        return '\n'.join(['%s:%s' % item for item in obj.__dict__.items()])

    def saveInExcel(self, resultfile, sheetname, resultrow):
        ws = resultfile.create_sheet(sheetname)
        for row in resultrow:
            ws.append(row)

    def bindvalue(self, first_index, first_second_index, first_table, second_index, second_table, second_values):
        tmpindex = [i[second_index] for i in second_table]
        result = []
        for row in first_table:
            line = []
            line.append(row[first_index])
            if (row[first_second_index] in tmpindex):
                secondline_index = tmpindex.index(row[first_second_index])
                line = line + [second_table[secondline_index][i] for i in second_values]
            result.append(line)
        return result
    
    def compareRecord(self, indexs, firstRec, secondRec):
        for i in range(len(firstRec)):
            if i in indexs:
                break
            if (firstRec[i] != secondRec[i]):
                return False
        return True

    def reverseSearch(self, value, index, Rec):
        #print('reverseSearch: ', value)
        for i in range(len(Rec)):
            #print('reverseSearch: searching', Rec[i])
            if (len(Rec[i]) <= index):
                continue
            #print('reverseSearch: searching', Rec[i][index])
            if (value == Rec[i][index]):
                #print('reverseSearch find: ', Rec[i][0])
                return i
        return -1

    def compareRule(self):
        cmgrule = self.cmgdata['PR']
        cmgpru = self.cmgdata['PRU']
        cmgcru = self.cmgdata['CRU']

        ngrule = self.ngdata['rule']

        ngrulestruct = [[i[j] for j in (0,3)] for i in ngrule]
        cmgrulestruct = self.bindvalue(0, 2, cmgrule, 0, cmgcru, [1])
        #print('compareRule cmgrulestruct: ', cmgrulestruct)

        cmgruleindex = [i[0] for i in cmgrule]
        #print('compareRule cmgruleindex: ', cmgruleindex)
        ngruleindex = [i[0] for i in ngrule]
        resultrow = []
        for ruledata in ngrulestruct:
            #print('compareRule: ', ruledata)
            newrulename = ruledata[0]
            if ruledata[0].split('_')[-1] == 'HeaderEnrich':
                newrulename = ruledata[0].replace('HeaderEnrich', 'L7')
            
            ruledataindex = -1
            if newrulename in cmgruleindex:
                ruledataindex = cmgruleindex.index(newrulename)
            elif ('PR_' + newrulename) in cmgruleindex:
                #print('compareRule ruledata: ', 'PR_' + newrulename)
                ruledataindex = cmgruleindex.index('PR_' + newrulename)
            if (ruledataindex >= 0):
                if not self.compareRecord([0], ruledata, cmgrulestruct[ruledataindex]):
                    resultrow.append(ruledata + cmgrulestruct[ruledataindex])
            else:
                if ('PR_' + newrulename + '_L7') in cmgruleindex:
                    #print('compareRule ruledata: ', 'PR_' + newrulename)
                    ruledataindex = cmgruleindex.index('PR_' + newrulename + '_L7')
                    resultrow.append(ruledata + cmgrulestruct[ruledataindex])
                else:
                    reverseIndex = self.reverseSearch(ruledata[1], 1, cmgrulestruct)
                    if reverseIndex >= 0:
                        resultrow.append(ruledata + cmgrulestruct[reverseIndex])
                    else:
                        resultrow.append(ruledata + [None, None])
        print('resultrow: ', resultrow)
        self.saveInExcel(self.resultfile, 'rule_rule', resultrow)


    def compareRuleBase(self):
        cmgrulebase = self.cmgdata['PRB']
        ngrulebase = self.ngdata['base']
        cmgrule = [i[1] for i in cmgrulebase]
        ngrule = [i[2] for i in ngrulebase]
        resultrow = []
        for ruledata in ngrulebase:
            newrulename = ruledata[2]
            if ruledata[2].split('_')[-1] == 'HeaderEnrich':
                newrulename = ruledata[2].replace('HeaderEnrich', 'L7')
            
            ruledataindex = -1
            if newrulename in cmgrule:
                ruledataindex = cmgrule.index(newrulename)
            elif ('PR_' + newrulename) in cmgrule:
                ruledataindex = cmgrule.index('PR_' + newrulename)
            if (ruledataindex >= 0):
                resultrow.append([ruledata[0], ruledata[2], cmgrulebase[ruledataindex][0],cmgrulebase[ruledataindex][1]])
            else:
                resultrow.append([ruledata[0], ruledata[2], None, None])
        self.saveInExcel(self.resultfile, 'rulebase_rule', resultrow)
        
    
    def closeAll(self):
        self.resultfile.save(self.resultfilename)
        
if __name__ == '__main__':
    filepath = os.path.split(os.path.realpath(__file__))[0]
    cmgfile = filepath + u'\\..\\tmp\\cmg内容计费读取v1.7 for sae07.xlsm'
    ngbasefile = filepath + u'\\..\\tmp\\sae07 base.xlsx'
    ngrulefile = filepath + u'\\..\\tmp\\sae07 rule.xlsx'
    ngfilterfile = filepath + u'\\..\\tmp\\sae07 filter.xlsx'
    resultfile = filepath + u'\\..\\tmp\\resultfile' + datetime.datetime.now().strftime('%Y%m%d%H%M%S') + '.xlsx'

    result = {
        'resultcode': '0',
        'resultdetail': ''
    }

    #print('cmg file: ', cmgfile)
    #print('ngbase file: ', ngbasefile)
    #print('ngrule file: ', ngrulefile)
    #print('ngfilter file: ', ngfilterfile)

    policyRuleCompare = PolicyRuleCompare(cmgfile, ngbasefile, ngrulefile, ngfilterfile, resultfile)
    policyRuleCompare.init()

    policyRuleCompare.compareRuleBase()
    policyRuleCompare.compareRule()

    policyRuleCompare.closeAll()
    
# test script
# python scripts\RTM_statis_new.py
	
