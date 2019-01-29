# -*- coding: utf-8 -*-

import sys
import os
import io
import json
import logging
import time
import datetime
import re
from ruamel.yaml import YAML

yaml=YAML(typ='safe')   # default, if not specfied, is 'rt' (round-trip)


import cx_Oracle as oracle
#import MySQLdb as mysql
import pyzabbix
import openpyxl

from MME_statis_new import *
from SAEGW_statis_new import *
from RTM_statis_new import *
from CMG_statis_new import *

class PM_ExcelFill:

    mmedb = None
    saegwdb = None
    cmgdb = None
    mmedbcursor = None
    saegwcursor = None
    cmgdbcursor = None
    ExcelConfigFileName = None
    Excel_Config = None
    Excel_Result = {}
    SaveFileName = ''
    rtm_statis = None
    workbook = None

    logging.basicConfig(filename='pm_excel_logger.log', level=logging.INFO)

    def __init__(self, param, runmode, excelconfig):
        self.param = param
        self.runmode = runmode
        self.filepath = os.path.split(os.path.realpath(__file__))[0]
        logging.info('filepath: ' + __file__ + self.filepath)
        if (self.runmode == 'test'):
            self.ExcelConfigFileName = self.filepath + '/config/' + excelconfig
        else:
            self.ExcelConfigFileName = self.filepath + '/config/' + runmode + '/' + excelconfig
        logging.info('Excel_Config_filename : ' + self.ExcelConfigFileName)
        logging.info('Excel Config File: ' + self.ExcelConfigFileName)

    def init(self):
        try:
            with open(self.ExcelConfigFileName, 'r') as load_f:
                ExcelConfigFileExt = self.ExcelConfigFileName.split('.') 
                if (ExcelConfigFileExt[-1] == 'json'):
                    self.Excel_Config = json.load(load_f)
                elif (ExcelConfigFileExt[-1] == 'yml'):
                    self.Excel_Config = yaml.load(load_f)
            #print('Excel_Config: ', self.Excel_Config)
            self.SaveFileName = self.Excel_Config['EXCEL_FILENAME']
            # MMEDB
            (dbuser,dbpasswd,dburl,dburlport,db_dbname) = getdbconfig(self.runmode, "mmedb")
            #print("mmedb parameter: ",dbuser,dbpasswd,dburl,dburlport,db_dbname)
            self.mmedb = oracle.connect(dbuser, dbpasswd, dburl)
            self.mmedbcursor=self.mmedb.cursor()
            # SAEGWDB
            (dbuser,dbpasswd,dburl,dburlport,db_dbname)=getdbconfig(self.runmode, "saegwdb")
            #print "saegwdb parameter: ", dbuser,dbpasswd,dburl,dburlport,db_dbname
            self.saegwdb = oracle.connect(dbuser, dbpasswd, dburl)
            self.saegwdbcursor=self.saegwdb.cursor()
            # CMGDB
            (dbuser,dbpasswd,dburl,dburlport,db_dbname)=getdbconfig(self.runmode, "cmgdb")
            #print "cmgdb parameter: ", dbuser,dbpasswd,dburl,dburlport,db_dbname
            self.cmgdb = oracle.connect(dbuser, dbpasswd, dburl)
            self.cmgdbcursor=self.cmgdb.cursor()

            if hasattr(self.param, 'netype') and self.param.netype == 'mme':
                self.SaveFileName += ('_' + self.param['ne'] or 'SHMME03BNK')
                self.param.isMME = 1
                self.param.isSAEGW = 0
                self.param.isCMG = 0
                self.param.selectmmesgsn = self.param['ne'] or 'SHMME03BNK'
                
            if hasattr(self.param, 'netype') and self.param.netype == 'saegw' :
                self.SaveFileName += ('_' + self.param['ne'] or "SHSAEGW03BNK")
                self.param.isMME = 0
                self.param.isSAEGW = 1
                self.param.isCMG = 0
                self.param.selectsaegwggsn = self.param['ne'] or 'SHSAEGW03BNK'
            
            if hasattr(self.param, 'netype') and self.param.netype == 'cmg' :
                self.SaveFileName += ('_' + self.param['ne'] or "SHSAEGW37BNK")
                self.param.isMME = 0
                self.param.isSAEGW = 0
                self.param.isCMG = 1
                self.param.selectcmg = self.param['ne'] or 'SHSAEGW37BNK'
            
            #print "before selectrtm"
            if hasattr(self.param, 'selectrtm') and self.param.selectrtm:
                (dbuser,dbpasswd,dburl,dburlport,db_dbname)=getdbconfig(self.runmode, "rtmdb")
                self.rtm_statis = RTM_Statis()
                if (self.runmode == 'test'):
                    result = self.rtm_statis.rtm_conn_win(dburl, dbuser, dbpasswd)
                else:
                    result = self.rtm_statis.rtm_conn_linux(dburl, dbuser, dbpasswd)
                if (result['resultcode'] == 0):
                    logging.info('self.rtm_statis init failed ' + result['result'])
                    raise Exception('initiate zabbix failed: %s' % result['result'])
            #print "after selectrtm"
            if(self.Excel_Config.has_key('SQLCUSTOM')):
                self.param['SQLCUSTOM'] = self.Excel_Config['SQLCUSTOM']
            logging.info('Excel_Config json parse finished')
            logging.info('Excel_Config EXCEL MODEL NAME: ' + self.Excel_Config['EXCEL_MODEL'])
            logging.info('Excel_Config EXCEL_FILENAME : ' + self.SaveFileName)
            self.workbook = openpyxl.load_workbook(self.Excel_Config['EXCEL_MODEL'])
            logging.info('openpyxl load workbook finished')
            return self.make_return(1, 'init ok')
        except Exception as e:
            logging.info('PM_ExcelFill initialize failed : %s' % str(e))
            return self.make_return(0, 'PM_ExcelFill initialize failed : %s' % str(e))

    def make_return(self, resultcode, result):
        return {'resultcode': resultcode, 'resultdetail': result}

    def namewithparam(self, name):
        logging.info('namewithparam origin name: ' + name)
        nameparams = re.findall('\$\{[\w\d]+\}', name)
        if (len(nameparams) > 0):
            for nameparam in nameparams:
                nameparamitem = re.search('([\w\d]+)', nameparam).group()
                nameparamstr = ''
                if (hasattr(self.param, nameparamitem)):
                    nameparamstr = self.param[nameparamitem]
                elif (nameparamitem in self.param['extraparams'].keys()):
                    nameparamstr = self.param['extraparams'][nameparamitem]
                name = name.replace(nameparam, nameparamstr)
        logging.info('namewithparam after name: ' + name)
        return name

    def excel_fill(self):
        try:        
            for sheet in self.Excel_Config['SHEETS']:
                if sheet.has_key('runCondition'):
                    logging.info(sheet['runCondition'] + ' : ' + str(self.param[sheet['runCondition']]))
                    logging.info('param has attr: ' + sheet['runCondition'] + ' ' + str(hasattr(self.param, sheet['runCondition'])) + ' ' + str(self.param[sheet['runCondition']]))
                    if hasattr(self.param, sheet['runCondition']) and not(self.param[sheet['runCondition']] == 1):
                        continue
                oldsheetname = sheet.has_key('SHEETNAME') and sheet['SHEETNAME']
                sheetname = oldsheetname
                if (sheet.has_key('SHEET_ORIGIN_NAME')):
                    logging.info('excel_fill sheet begin: ' + sheet['SHEET_ORIGIN_NAME'])
                    oldsheetname = sheet['SHEET_ORIGIN_NAME']
                if (sheet.has_key('SHEET_AFTER_NAME')):
                    logging.info('excel_fill sheet after: ' + sheet['SHEET_AFTER_NAME'])
                    sheetname = sheet['SHEET_AFTER_NAME']
                sheetname = self.namewithparam(sheetname)
                logging.info('sheetname after param modify is: ' + sheetname)
                ws = self.workbook[oldsheetname]
                ws.title = sheetname
                self.saveExcelSheet(ws, sheet)
                logging.info('excel_fill sheet over: ' + sheetname) 
            if (self.runmode == 'test'):
                site_config_filename = self.filepath + '/config/api_options.json'
            else:
                site_config_filename = self.filepath + '/config/' + self.runmode + '/api_options.json' 
            logging.info('site_config_filename  : ' + site_config_filename)
            site_config = json.load(open(site_config_filename, 'r'))
            
            #nowtime = datetime.datetime.now() 
            #self.SaveFileName = self.SaveFileName + '_' + nowtime.strftime("%Y%m%d") + nowtime.strftime("%H%M%S") + '.xlsx'
            self.SaveFileName = self.SaveFileName + '_' + self.param.maketime + '.xlsx'
            realfilename = site_config['download_dir'] + self.SaveFileName 
            self.workbook.save(realfilename)
            return self.make_return(1, self.Excel_Config['EXCEL_DOWNLOAD_URL'] + self.SaveFileName)
        except Exception as e:
            logging.error("Error PM_Excelfill: %s" % e)
            return self.make_return(0, "Error PM_Excelfill: %s" % e)
        
    def closeAll(self):
        if (self.mmedb != None):
            self.mmedb.close
        if (self.saegwdb != None):
            self.saegwdb.close
        if (self.cmgdb != None):
            self.cmgdb.close

    def prn_obj(self, obj):
        return '\n'.join(['%s:%s' % item for item in obj.__dict__.items()])


    def avg_cacul(self, algo, rows, algoindex):
        sum = 0
        for row in rows:
            sum += float(row[int(algoindex)-1])
        if(len(rows)>0):
            return str(sum/len(rows))
        return 0

    def max_cacul(self, algo, rows, algoindex):
        max = 0
        for row in rows:
            if(float(row[int(algoindex)-1]) > max):
                max = float(row[int(algoindex)-1])
        return str(max)

    def min_cacul(self, algo, rows, algoindex):
        min = 0
        for row in rows:
            if(float(row[int(algoindex)-1]) < min):
                min = float(row[int(algoindex)-1])
        return str(min)

    def sum_cacul(self, algo, rows, algoindex):
        sum = 0
        for row in rows:
            sum += float(row[int(algoindex)-1])
        return str(sum)


    def algo_cacul(self, algo, rows, algoindex):
        ret = '0'
        if (algo == 'avg'):
            ret = self.avg_cacul(algo, rows, algoindex)
        elif (algo == 'max'):
            ret = self.max_cacul(alog, rows, algoindex)
        elif (algo == 'min'):
            ret = self.min_cacul(alog, rows, algoindex)
        elif (algo == 'sum'):
            ret = self.sum_cacul(alog, rows, algoindex)
        
        return ret

    def caculate(self, rows, sqlextrastruct):
        tmprow = []
        if (sqlextrastruct.has_key('valuefilter')):
            itemindex = sqlextrastruct['valuefilter']['filter_index']
            filter = sqlextrastruct['valuefilter']['filter_regex']
            #print 'caculate filter: ', filter
            for row in rows:
                #print('caculate filter row: ', str(row))
                if(re.match(filter, row[int(itemindex)-1])):
                    tmprow.append(row)
        else:
            tmprow = rows
        #print('caculate after filter: ', tmprow)
        if (sqlextrastruct.has_key('valuealgo')):
            algoindex = sqlextrastruct['valuealgo']['algo_index']
            algo = sqlextrastruct['valuealgo']['algo']
            ret = self.algo_cacul(algo, tmprow, algoindex)
            #print('caculate after algo: ' + str(ret))
            return ret
        else:
            return tmprow

    # this func for generate result keypart. 
    # if valuestruct is same(except sql_selectitem_index)
    # then we can use the same result 
    def sqlkeypart(self, sqlfunc, valuestruct):
        tmpvaluestruct = valuestruct.copy()
        tmpvaluestruct['sql_selectitem_index'] = ''
        sqlkeypartstring = sqlfunc + json.dumps(tmpvaluestruct)
        return sqlkeypartstring

    def getsqlinfo(self, dbcursor, valuestruct, api_sql_function, kpi_report_result):
        try:
            sqlfunc = valuestruct['sql_function']
            #print('getsqlinfo : ', sqlfunc)
            sqlitemindex = valuestruct['sql_selectitem_index']
            #print('getsqlinfo : ', sqlitemindex)
            kpi_function = api_sql_function[sqlfunc]['func']
            #print('getsqlinfo : ', kpi_function)
            if (valuestruct.has_key('netelement')):
                sqlnetelement = valuestruct['netelement'].split(':')
                sqlnetype = sqlnetelement[0]
                sqlnename = sqlnetelement[1]
                self.param[sqlnetype] = sqlnename
            # for kindof reportsuite func, we need a sql template name 
            if (valuestruct.has_key('sql_template')):
                self.param['sql_template'] = valuestruct['sql_template']

            if (kpi_report_result.has_key(self.sqlkeypart(sqlfunc, valuestruct))):
                if (valuestruct.has_key("sql_extra")):
                    #print('getsqlinfo valuestruct has key sql_extra')
                    result = []
                    kpiresult = self.caculate(kpi_report_result[self.sqlkeypart(sqlfunc, valuestruct)], valuestruct["sql_extra"])
                    for kpiresultitem in kpiresult:
                        result.append(str(kpiresultitem[int(sqlitemindex) - 1]))
                    return result
                else:    
                    result = []
                    kpiresult = kpi_report_result[self.sqlkeypart(sqlfunc, valuestruct)]
                    for kpiresultitem in kpiresult:
                        result.append(str(kpiresultitem[int(sqlitemindex) - 1]))
                    return result
            else:
                title,row=kpi_function(sqlfunc, dbcursor, self.param)
                #print 'kpi_function param: ', self.param
                if title[0]!='error' and len(row)>0:
                    #print('getsqlinfo ' + sqlfunc + ' ' + str(row[0]))
                    kpi_report_result[self.sqlkeypart(sqlfunc, valuestruct)] = row
                    #print('getsqlinfo valuestruct has key sql_extra ? ', valuestruct.has_key("sql_extra"))
                    if (valuestruct.has_key("sql_extra")):
                        #print('getsqlinfo valuestruct has key sql_extra')
                        result = []
                        kpiresult = self.caculate(row, valuestruct["sql_extra"])
                        for kpiresultitem in kpiresult:
                            result.append(str(kpiresultitem[int(sqlitemindex) - 1]))
                        return result
                    else:
                        result = []
                        kpiresult = kpi_report_result[self.sqlkeypart(sqlfunc, valuestruct)]
                        for kpiresultitem in kpiresult:
                            result.append(str(kpiresultitem[int(sqlitemindex) - 1]))
                        return result
                    #return str(row[0][int(sqlitemindex) - 1])
                else:
                    logging.error("getsqlinfo %s with Error : %s" % (sqlfunc, title[1]))
                    return " "

        except Exception, e:
            logging.error('Exception : ' + str(e))
            return "getsqlinfo Exception: %s" % str(e)

    def getData(self, outputformat, dbcursor, values, api_sql_function, kpi_report_result):
        logging.info('getData...')
        if (outputformat['type'] == "string"):
            return outputformat['value']
        elif(outputformat['type'] == "data"):
            if (values[int(outputformat['value']) - 1] != None):
                datasource = values[int(outputformat['value']) - 1]['datasource']
                if (datasource == 'params'):
                    datavalue = values[int(outputformat['value']) - 1]['datavalue']
                    return param[datavalue]
                elif(datasource == "sql"):
                    datainfo = self.getsqlinfo(dbcursor, values[int(outputformat['value'])-1], api_sql_function, kpi_report_result)
                    logging.info('sql datainfo: %s' % datainfo)
                    return datainfo
                elif(datasource == 'evaldata'):
                    evalstring = values[int(outputformat['value']) - 1]['value']
                    logging.info('evalstring begin: ' + evalstring)
                    evalsubitems = re.findall('\$\{\d+\}', evalstring)
                    logging.info('evalsubitems : ' + str(evalsubitems))
                    evalstrings = []
                    if (len(evalsubitems) > 0):
                        for evalsubitem in evalsubitems:
                            evalsubitem_match = re.search('(\d+)', evalsubitem)
                            logging.info('evalsubitem_match : ' + evalsubitem_match.group())
                            if(evalsubitem_match):
                                evalsubitem_index = int(evalsubitem_match.group())
                                newoutputformat = {
                                    "type": "data",
                                    "value": evalsubitem_index
                                }
                                result = self.getData(newoutputformat, dbcursor, values, api_sql_function, kpi_report_result)
                                #logging.info('result: ' , result)
                                i = 0
                                for resultitem in result:
                                    if (len(evalstrings)>i):
                                        evalstrings[i] = evalstrings[i].replace(evalsubitem, resultitem)
                                    else:
                                        tmpevalstring = str(evalstring)
                                        tmpevalstring = tmpevalstring.replace(evalsubitem, resultitem)
                                        evalstrings.append(tmpevalstring)
                                    i = i + 1
                    try:
                        #print ('evalstring over: ' , evalstrings)
                        afterevalstrings = []
                        for evalstring in evalstrings:
                            #logging.info('evalstring: ', str(evalstring))
                            afterevalstrings.append(str(eval(evalstring)))
                        return afterevalstrings
                    except Exception as e:
                        logging.info('evalstring over Error: ' + str(e))
                        return str(e)


    def saveInExcel(self, datastr, KPI, ws):
        if (type(KPI['outputlocation']) == list):
            datastrlist = datastr.split(',')
            #print 'saveInExcel datastrlist: ', datastrlist
            for i in range(min(len(datastrlist), len(KPI['outputlocation']))): 
                ws[KPI['outputlocation'][i]] = datastrlist[i]
            #print 'saveInExcel over'
        else:
            ws[KPI['outputlocation']] = datastr

    def saveExcelSheet(self, ws, sheet):
        if 'RTM_KPI' in sheet.keys():
            self.saveExcelSheetRTMKPI(ws, sheet['RTM_KPI'])
        if 'MME_KPI' in sheet.keys():
            self.saveExcelSheetMMEKPI(ws, self.mmedbcursor, sheet['MME_KPI'])
        if 'SAEGW_KPI' in sheet.keys():
            self.saveExcelSheetSAEGWKPI(ws, self.saegwdbcursor, sheet['SAEGW_KPI'])
        if 'CMG_KPI' in sheet.keys():
            self.saveExcelSheetCMGKPI(ws, self.cmgdbcursor, sheet['CMG_KPI'])

    def saveExcelSheetRTMKPI(self, ws, kpiconfig):
        kpi_list = kpiconfig
        kpi_report_result = {}
        self.runRTMKPI(ws, kpi_list)
        
    def runRTMKPI(self, ws, kpi_list):    
        for kpi in kpi_list:
            #print 'runRTMKPI: ', kpi
            values = kpi['values']
            outputformats = kpi['outputformats']
            outputstring = ''
            for outputformat in outputformats:
                if (outputformat['type'] == 'string'):
                    outputstring += outputformat['value']
                elif values[int(outputformat['value']) - 1]['datasource'] == 'params':
                    datavalue = values[int(outputformat['value']) - 1]['datavalue']
                    if (hasattr(self.param, datavalue)):
                        outputstring += self.param[datavalue]
                    elif (datavalue in self.param['extraparams'].keys()):
                        outputstring += self.param['extraparams'][datavalue]
                elif values[int(outputformat['value']) - 1]['datasource'] == 'rtm':
                    valuesindex = int(outputformat['value']) - 1
                    result = self.getRTMKPI(values, valuesindex)
                    outputstring += ','.join(result)
                elif values[int(outputformat['value']) - 1]['datasource'] == 'evaldata':
                    evalstring = values[int(outputformat['value']) - 1]['value']
                    logging.info('evalstring begin: ' + evalstring)
                    evalsubitems = re.findall('\$\{\d+\}', evalstring)
                    logging.info('evalsubitems : %s' % evalsubitems)
                    evalstrings = []
                    if (len(evalsubitems) > 0):
                        for evalsubitem in evalsubitems:
                            evalsubitem_match = re.search('(\d+)', evalsubitem)
                            logging.info('evalsubitem_match : ' + evalsubitem_match.group())
                            if(evalsubitem_match):
                                evalsubitem_index = int(evalsubitem_match.group()) - 1
                                result = self.getRTMKPI(values, evalsubitem_index)
                                logging.info('result: ' , result)
                                i = 0
                                for resultitem in result:
                                    if (len(evalstrings)>i):
                                        evalstrings[i] = evalstrings[i].replace(evalsubitem, resultitem)
                                    else:
                                        tmpevalstring = str(evalstring)
                                        tmpevalstring = tmpevalstring.replace(evalsubitem, resultitem)
                                        evalstrings.append(tmpevalstring)
                                    i = i + 1
                    try:
                        #print ('evalstring over: ' , evalstrings)
                        afterevalstrings = []
                        for evalstring in evalstrings:
                            #logging.info('evalstring: ', str(evalstring))
                            afterevalstrings.append(str(eval(evalstring)))
                        return afterevalstrings
                    except Exception as e:
                        outputstring = outputstring + evalstring + str(e)
            self.saveInExcel(outputstring, kpi, ws)

    def getRTMKPI(self, values, valuesindex):
        hostname = values[valuesindex]['host']
        itemname = values[valuesindex]['itemname']
        valueindex = values[valuesindex]['valueindex']
        resultvalues = self.rtm_statis.rtm_get_value(hostname, itemname, self.param.startdatetime, self.param.stopdatetime)
        #print('rtm resultvalues: ' , resultvalues, resultvalues['result'])
        result = []
        if (resultvalues['resultcode'] == 1 ):
            for resultvalue in resultvalues['result']:
                #print('resultvalues["result"]: ', resultvalue)
                if (valueindex == 'min'):
                    result.append(str(round(float(resultvalue['value_min']),2))) 
                if (valueindex == 'avg'):
                    result.append(str(round(float(resultvalue['value_avg']),2)))
                else:
                    result.append(str(round(float(resultvalue['value_max']),2)))
        #print 'getRTMKPI result: ', result
        return result

    def saveExcelSheetSAEGWKPI(self, ws, dbcursor, kpiconfig):
        logging.info('saveExcelSheetSAEGWKPI...')
        api_sql_function = saegw_api_sql_function
        kpi_list = kpiconfig
        kpi_report_result = {}
        self.runOSSKPI(ws, kpi_list, dbcursor, api_sql_function, kpi_report_result)

    def saveExcelSheetMMEKPI(self, ws, dbcursor, kpiconfig):
        api_sql_function = mme_api_sql_function
        kpi_list = kpiconfig
        kpi_report_result = {}
        self.runOSSKPI(ws, kpi_list, dbcursor, api_sql_function, kpi_report_result)

    def saveExcelSheetCMGKPI(self, ws, dbcursor, kpiconfig):
        api_sql_function = cmg_api_sql_function
        kpi_list = kpiconfig
        kpi_report_result = {}
        self.runOSSKPI(ws, kpi_list, dbcursor, api_sql_function, kpi_report_result)
        
    def runOSSKPI(self, ws, kpi_list, dbcursor, api_sql_function, kpi_report_result):    
        for kpi in kpi_list:
            #logging.info('kpi : ' + kpi)
            logging.info('runOSSKPI: ' + json.dumps(kpi))
            values = kpi['values']
            outputformats = kpi['outputformats']
            outputstring = ''
            for outputformat in outputformats:
                if (outputformat['type'] == 'string'):
                    outputstring += outputformat['value']
                elif values[int(outputformat['value']) - 1]['datasource'] == 'params':
                    datavalue = values[int(outputformat['value']) - 1]['datavalue']
                    if (hasattr(self.param, datavalue)):
                        outputstring += self.param[datavalue]
                    elif (datavalue in self.param['extraparams'].keys()):
                        outputstring += self.param['extraparams'][datavalue]
                else:
                    outputstring += ','.join(self.getData(outputformat, dbcursor, values, api_sql_function, kpi_report_result))
            #print "runOSSKPI outputstring: ", outputstring
            self.saveInExcel(outputstring, kpi, ws)

def getforminfo(params, formparams):
    # Get data from formparams
    configs = json.loads(formparams)
    logging.info("formparams : " + formparams)
    #print configs
    for key, item in configs.items():
        #print key," : ",item
        params.__dict__[key] = item
    if (not configs.has_key('selectperiod')):
        params.selectperiod = '60'
    if (params.selectperiodtype == None):
        params.selectperiodtype = 'continue'
    
    
def paramsdate_fix():
    # adjust per environment in test or prod
    stopdate = time.strftime('%Y/%m/%d',time.localtime(time.time()))
    startdate = time.strftime('%Y/%m/%d',time.localtime(time.time()-3600))
    curtime = time.strftime('%H',time.localtime(time.time()))
    curtime = curtime+":00"
    pretime = time.strftime('%H',time.localtime(time.time()-3600))
    pretimr = pretime+":00"
    
    # for test , I only have some days data, so I will adjust the stopdate
    startdate = '2017/02/11' 
    if (curtime < '01:00'):
        stopdate = '2017/02/12'
    else:
        stopdate = '2017/02/11'

    param.starttime=pretime
    param.stoptime=curtime
    param.startdate=startdate
    param.stopdate=stopdate
        
if __name__ == '__main__':

    logging.info('query time : ' + time.strftime('%Y/%m/%d %H:%M:%S',time.localtime(time.time())))
    formparams = None
    runmode = 'test'
    if ( len(sys.argv) ) > 2 :
        logging.info("\t run mode : " + str(sys.argv[1]))
        logging.info("\t excel config : " + str(sys.argv[2]))
        logging.info("\t params : " + str(sys.argv[3])) 
        runmode = sys.argv[1]
        excelconfig = sys.argv[2]
        formparams = sys.argv[3]
        logging.info("\t run mode : " + runmode) 
        logging.info("\t excel config : " + excelconfig) 
        logging.info("\t formparams : " + formparams) 
         
    param = PmSqlParam()

    if (not formparams == None):
        try:
            getforminfo(param, formparams)
            logging.info('\tparams info : ' + prn_obj(param))
        except Exception as e:
            logging.error('error in param get: %s' % e )
    else:
        logging.info('form params is None.')
    
    # for excel , in these param
    # we need netype(mme or saegw), ne(SHMME03BNK), time(2018-06-30T13:03:51.155Z)

    #currtime = datetime.datetime.strptime(param.time, "%Y-%m-%dT%H:%M:%S.%fZ")
    if not hasattr(param, 'time') or param['time'] == None:
        param['time'] = time.time()
    timeArray = time.localtime(float(param.time))
    param_time = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
    currtime = datetime.datetime.strptime(param_time, "%Y-%m-%d %H:%M:%S")
    currtime = datetime.datetime(currtime.year, currtime.month, currtime.day, currtime.hour, 0)
    currtime = currtime + datetime.timedelta(minutes=-1)
    pretime = currtime + datetime.timedelta(minutes=-1*59)
    
    param['startdatetime'] = time.mktime(pretime.timetuple())
    param['stopdatetime'] = time.mktime(currtime.timetuple())

    param.startdate = pretime.strftime("%Y/%m/%d")
    param.stopdate = currtime.strftime("%Y/%m/%d")
    param.starttime = pretime.strftime("%H:%M")
    param.stoptime = currtime.strftime("%H:%M")
    param.maketime = time.strftime("%Y%m%d%H%M%S", timeArray)
    param['extraparams']['year'] = currtime.strftime('%Y')
    param['extraparams']['month'] = currtime.strftime('%m')
    param['extraparams']['day'] = currtime.strftime('%d')
    param['extraparams']['hour'] = currtime.strftime('%H')
    param['extraparams']['minute'] = currtime.strftime('%M')
    logging.info('param: \n%s' % param)

    filepath = os.path.split(os.path.realpath(__file__))[0]
    logging.info('main filepath: ' + filepath)
    pm_excelfill = PM_ExcelFill(param, runmode, excelconfig)

    fill_result = pm_excelfill.init()
    if fill_result['resultcode'] == 0:
        logging.info('PM_ExcelFill initialize failed : %s' % fill_result['resultdetail'])
        #sys.exit(1)
    else:
        fill_result = pm_excelfill.excel_fill()
        if fill_result['resultcode'] == 0:
            logging.info('PM_ExcelFill fill failed : %s' % fill_result['resultdetail'])
            #sys.exit(1)
    
    pm_excelfill.closeAll()
    #print(json.dumps(fill_result['result']))
    print(json.dumps(fill_result))

# test script
# time : 2019, 1, 26, 2, 50, 49
# python PM_ExcelFill_v2.py test EPCPSCore_2019SF.yml {\"selectsaegwggsn\":\"SHSAEGW03BNK\",\"selectrtm\":\"true\",\"time\":\"1548499849"}
	
